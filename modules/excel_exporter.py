from __future__ import annotations

import io
from typing import List, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from modules.data_processor import Quote, find_best_price_per_item


_C_DARK_BLUE = "1F4E79"
_C_MID_BLUE = "2E75B6"
_C_LIGHT_BLUE = "D6E4F0"
_C_ALT_ROW = "EBF3FB"
_C_BEST_BG = "C6EFCE"
_C_BEST_FONT = "276221"
_C_WORST_BG = "FFCCCC"
_C_WORST_FONT = "9C0006"
_C_TOTAL_BG = "FFF2CC"
_C_TOTAL_FONT = "7F6000"
_C_WHITE = "FFFFFF"
_C_BORDER = "BFBFBF"


def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _thin_border() -> Border:
    s = Side(style="thin", color=_C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _medium_border() -> Border:
    m = Side(style="medium", color=_C_DARK_BLUE)
    return Border(left=m, right=m, top=m, bottom=m)


def _style_header(
    cell,
    bg: str = _C_DARK_BLUE,
    fg: str = _C_WHITE,
    size: int = 10,
    bold: bool = True,
    wrap: bool = True,
    align: str = "center",
) -> None:
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin_border()


def _style_data(
    cell,
    bold: bool = False,
    number_format: Optional[str] = None,
    bg: Optional[str] = None,
    fg: str = "000000",
    align: str = "left",
    size: int = 10,
    wrap: bool = True,
) -> None:
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin_border()
    if bg:
        cell.fill = _fill(bg)
    if number_format:
        cell.number_format = number_format


def _set_col_widths(ws, widths: List[int]) -> None:
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _auto_width(ws, min_w: int = 10, max_w: int = 55) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                lines = str(cell.value).split("\n")
                max_len = max(max_len, max(len(ln) for ln in lines))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, min_w), max_w)


def _write_title_band(ws, text: str, ncols: int, row: int = 1, height: int = 30) -> None:
    end_col = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, size=14, color=_C_WHITE)
    cell.fill = _fill(_C_DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _medium_border()
    ws.row_dimensions[row].height = height


def _write_provider_block(ws, quote: Quote, start_row: int) -> int:
    row = start_row
    ws.merge_cells(f"A{row}:F{row}")
    section_cell = ws[f"A{row}"]
    section_cell.value = "INFORMACIÓN DEL PROVEEDOR"
    _style_header(section_cell, bg=_C_MID_BLUE, size=11)
    ws.row_dimensions[row].height = 20
    row += 1

    fields = [
        ("Proveedor / Razón Social:", quote.proveedor),
        ("RUC:", quote.ruc),
        ("Dirección:", quote.direccion),
        ("Teléfono:", quote.telefono),
        ("N° Cotización:", quote.numero_cotizacion),
        ("Fecha:", quote.fecha),
        ("Vigencia:", quote.vigencia),
        ("Observaciones:", quote.observaciones),
    ]
    for label, value in fields:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name="Calibri", bold=True, size=10, color=_C_DARK_BLUE)
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.fill = _fill(_C_LIGHT_BLUE)
        label_cell.border = _thin_border()

        ws.merge_cells(f"B{row}:F{row}")
        val_cell = ws[f"B{row}"]
        val_cell.value = value
        val_cell.font = Font(name="Calibri", size=10)
        val_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        val_cell.border = _thin_border()
        ws.row_dimensions[row].height = 16
        row += 1
    return row


def _write_items_block(ws, quote: Quote, start_row: int) -> int:
    row = start_row
    ws.merge_cells(f"A{row}:F{row}")
    section_cell = ws[f"A{row}"]
    section_cell.value = "DETALLE DE ÍTEMS"
    _style_header(section_cell, bg=_C_MID_BLUE, size=11)
    ws.row_dimensions[row].height = 20
    row += 1

    col_headers = [
        "Código CPC/UNSPSC", "Descripción", "Unidad",
        "Cantidad", "Precio Unitario", "Precio Total"
    ]
    col_widths_items = [18, 42, 12, 12, 16, 16]
    _set_col_widths(ws, col_widths_items)

    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        _style_header(cell, bg=_C_DARK_BLUE, size=10)
    ws.row_dimensions[row].height = 18
    row += 1

    for idx, item in enumerate(quote.items):
        alt = _C_ALT_ROW if idx % 2 == 0 else None
        data = [
            item.codigo_cpc,
            item.descripcion,
            item.unidad,
            item.cantidad,
            item.precio_unitario,
            item.precio_total,
        ]
        for ci, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=ci, value=value)
            if ci == 4:
                _style_data(cell, bg=alt, number_format="#,##0.00", align="center")
            elif ci in (5, 6):
                _style_data(cell, bg=alt, number_format='"$"#,##0.00', align="right")
            else:
                _style_data(cell, bg=alt, align="left")
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1
    totals_data = [
        ("SUBTOTAL:", quote.subtotal),
        (f"IVA ({quote.iva_porcentaje}%):", quote.iva),
        ("TOTAL:", quote.total),
    ]
    for label, amount in totals_data:
        for ci in range(1, 7):
            cell = ws.cell(row=row, column=ci)
            if ci == 5:
                cell.value = label
                cell.font = Font(name="Calibri", bold=True, size=10, color=_C_TOTAL_FONT)
                cell.fill = _fill(_C_TOTAL_BG)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _thin_border()
            elif ci == 6:
                cell.value = amount
                cell.font = Font(name="Calibri", bold=True, size=11, color=_C_TOTAL_FONT)
                cell.fill = _fill(_C_TOTAL_BG)
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _thin_border()
            else:
                _style_data(cell)
        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _write_quote_sheet(wb: Workbook, quote: Quote) -> None:
    raw = quote.proveedor[:28].strip() if quote.proveedor else "Proveedor"
    safe = "".join(
        c for c in raw if c not in ("/", "\\", "*", "?", "[", "]", ":", "'")
    ).strip() or f"Prov_{quote.id[:6]}"

    existing = [s.title for s in wb.worksheets]
    base, counter = safe, 1
    while safe in existing:
        safe = f"{base[:24]}_{counter}"
        counter += 1

    ws = wb.create_sheet(title=safe)

    _write_title_band(
        ws,
        f"COTIZACIÓN — {quote.proveedor.upper()}",
        ncols=6,
        row=1,
        height=28,
    )

    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 6

    current_row = _write_provider_block(ws, quote, start_row=3)
    current_row += 1
    _write_items_block(ws, quote, start_row=current_row)

    ws.sheet_view.showGridLines = True
    ws.freeze_panes = ws.cell(row=3, column=1)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.print_title_rows = "1:2"


def _write_comparison_sheet(wb: Workbook, quotes: List[Quote]) -> None:
    ws = wb.create_sheet(title="Comparativo", index=0)
    best = find_best_price_per_item(quotes)

    total_data_cols = 1 + len(quotes) * 2
    _write_title_band(ws, "CUADRO COMPARATIVO DE COTIZACIONES — SERCOP/SOCE", total_data_cols, row=1)

    ws.merge_cells("A2:A2")
    ws.row_dimensions[2].height = 6

    header_row = 3
    desc_cell = ws.cell(row=header_row, column=1, value="Descripción del Ítem")
    _style_header(desc_cell, bg=_C_DARK_BLUE, size=10)
    ws.column_dimensions["A"].width = 44

    for qi, q in enumerate(quotes):
        base_col = 2 + qi * 2
        ws.merge_cells(
            start_row=header_row, start_column=base_col,
            end_row=header_row, end_column=base_col + 1
        )
        prov_cell = ws.cell(row=header_row, column=base_col, value=q.proveedor)
        _style_header(prov_cell, bg=_C_MID_BLUE, size=10)
        ws.row_dimensions[header_row].height = 28

    sub_row = header_row + 1
    sub_desc = ws.cell(row=sub_row, column=1, value="")
    _style_header(sub_desc, bg=_C_DARK_BLUE, size=9)

    for qi in range(len(quotes)):
        base_col = 2 + qi * 2
        pu = ws.cell(row=sub_row, column=base_col, value="Precio Unitario")
        _style_header(pu, bg=_C_DARK_BLUE, size=9)
        pt = ws.cell(row=sub_row, column=base_col + 1, value="Precio Total")
        _style_header(pt, bg=_C_DARK_BLUE, size=9)
        ws.column_dimensions[get_column_letter(base_col)].width = 15
        ws.column_dimensions[get_column_letter(base_col + 1)].width = 15
    ws.row_dimensions[sub_row].height = 18

    all_descriptions: List[str] = []
    for q in quotes:
        for item in q.items:
            d = item.descripcion.strip()
            if d and d not in all_descriptions:
                all_descriptions.append(d)

    data_start = sub_row + 1
    for idx, desc in enumerate(all_descriptions):
        row_num = data_start + idx
        alt = _C_ALT_ROW if idx % 2 == 0 else None

        desc_cell = ws.cell(row=row_num, column=1, value=desc)
        _style_data(desc_cell, bg=alt, align="left")

        best_provider = best.get(desc)

        all_prices = []
        for q in quotes:
            m = next((it for it in q.items if it.descripcion.strip() == desc), None)
            if m and m.precio_unitario > 0:
                all_prices.append(m.precio_unitario)
        min_price = min(all_prices) if all_prices else None

        for qi, q in enumerate(quotes):
            base_col = 2 + qi * 2
            matched = next(
                (it for it in q.items if it.descripcion.strip() == desc), None
            )
            pu_cell = ws.cell(row=row_num, column=base_col)
            pt_cell = ws.cell(row=row_num, column=base_col + 1)

            if matched and (matched.precio_unitario > 0 or matched.precio_total > 0):
                is_best = best_provider == q.proveedor
                is_worst = (
                    not is_best
                    and len(all_prices) > 1
                    and min_price is not None
                    and matched.precio_unitario > min_price
                )

                if is_best:
                    item_bg = _C_BEST_BG
                    item_fg = _C_BEST_FONT
                    bold = True
                elif is_worst:
                    item_bg = _C_WORST_BG
                    item_fg = _C_WORST_FONT
                    bold = False
                else:
                    item_bg = alt
                    item_fg = "000000"
                    bold = False

                pu_cell.value = matched.precio_unitario
                _style_data(
                    pu_cell, bold=bold, number_format='"$"#,##0.00',
                    bg=item_bg, fg=item_fg, align="right"
                )
                pt_cell.value = matched.precio_total
                _style_data(
                    pt_cell, bold=bold, number_format='"$"#,##0.00',
                    bg=item_bg, fg=item_fg, align="right"
                )
            else:
                pu_cell.value = "N/D"
                pt_cell.value = "N/D"
                _style_data(pu_cell, bg=_C_WORST_BG, fg=_C_WORST_FONT, align="center")
                _style_data(pt_cell, bg=_C_WORST_BG, fg=_C_WORST_FONT, align="center")

        ws.row_dimensions[row_num].height = 15

    total_row = data_start + len(all_descriptions)
    ws.row_dimensions[total_row].height = 22

    total_label = ws.cell(row=total_row, column=1, value="TOTAL COTIZACIÓN (con IVA 12%)")
    total_label.font = Font(name="Calibri", bold=True, size=11, color=_C_TOTAL_FONT)
    total_label.fill = _fill(_C_TOTAL_BG)
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    total_label.border = _thin_border()

    best_total_provider = best.get("__TOTAL__")
    all_totals = [q.total for q in quotes if q.total > 0]
    min_total = min(all_totals) if all_totals else None

    for qi, q in enumerate(quotes):
        base_col = 2 + qi * 2
        ws.merge_cells(
            start_row=total_row, start_column=base_col,
            end_row=total_row, end_column=base_col + 1
        )
        total_cell = ws.cell(row=total_row, column=base_col)
        total_cell.value = q.total
        total_cell.number_format = '"$"#,##0.00'
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell.border = _thin_border()

        is_best_total = best_total_provider == q.proveedor
        is_worst_total = (
            not is_best_total
            and len(all_totals) > 1
            and min_total is not None
            and q.total > min_total
        )

        if is_best_total:
            total_cell.font = Font(name="Calibri", bold=True, size=12, color=_C_BEST_FONT)
            total_cell.fill = _fill(_C_BEST_BG)
        elif is_worst_total:
            total_cell.font = Font(name="Calibri", bold=True, size=11, color=_C_WORST_FONT)
            total_cell.fill = _fill(_C_WORST_BG)
        else:
            total_cell.font = Font(name="Calibri", bold=True, size=11, color=_C_TOTAL_FONT)
            total_cell.fill = _fill(_C_TOTAL_BG)

    legend_row = total_row + 2
    ws.row_dimensions[legend_row].height = 14
    ws.cell(row=legend_row, column=1, value="Leyenda:").font = Font(
        name="Calibri", bold=True, size=9
    )
    best_legend = ws.cell(row=legend_row + 1, column=1, value="  Mejor precio")
    best_legend.fill = _fill(_C_BEST_BG)
    best_legend.font = Font(name="Calibri", bold=True, color=_C_BEST_FONT, size=9)
    best_legend.border = _thin_border()
    ws.row_dimensions[legend_row + 1].height = 13

    worst_legend = ws.cell(row=legend_row + 2, column=1, value="  Precio más alto")
    worst_legend.fill = _fill(_C_WORST_BG)
    worst_legend.font = Font(name="Calibri", bold=False, color=_C_WORST_FONT, size=9)
    worst_legend.border = _thin_border()
    ws.row_dimensions[legend_row + 2].height = 13

    ws.freeze_panes = ws.cell(row=data_start, column=2)
    ws.sheet_view.showGridLines = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True


def _write_summary_sheet(wb: Workbook, quotes: List[Quote]) -> None:
    ws = wb.create_sheet(title="Resumen", index=1)

    ncols = 9
    _write_title_band(
        ws,
        "RESUMEN GENERAL DE COTIZACIONES — CONTRATACIÓN PÚBLICA (SERCOP/SOCE)",
        ncols=ncols,
        row=1,
    )
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 6

    headers = [
        "N°", "Proveedor / Razón Social", "RUC", "N° Cotización",
        "Fecha", "Vigencia", "Subtotal", "IVA (12%)", "Total"
    ]
    col_widths_sum = [6, 36, 16, 18, 14, 20, 16, 14, 16]
    _set_col_widths(ws, col_widths_sum)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        _style_header(cell, bg=_C_MID_BLUE, size=10)
    ws.row_dimensions[3].height = 20

    best_total_map = find_best_price_per_item(quotes)
    best_total_prov = best_total_map.get("__TOTAL__")
    all_totals = [q.total for q in quotes if q.total > 0]
    min_total = min(all_totals) if all_totals else None

    for idx, q in enumerate(quotes):
        row_num = 4 + idx
        alt = _C_ALT_ROW if idx % 2 == 0 else None
        is_best = best_total_prov == q.proveedor
        is_worst = (
            not is_best
            and len(all_totals) > 1
            and min_total is not None
            and q.total > min_total
        )

        data_bg = _C_BEST_BG if is_best else (_C_WORST_BG if is_worst else alt)
        num_fg = _C_BEST_FONT if is_best else (_C_WORST_FONT if is_worst else "000000")

        row_data = [
            idx + 1,
            q.proveedor,
            q.ruc,
            q.numero_cotizacion,
            q.fecha,
            q.vigencia,
            q.subtotal,
            q.iva,
            q.total,
        ]
        for ci, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=ci, value=value)
            if ci in (7, 8, 9):
                _style_data(
                    cell, bold=is_best, number_format='"$"#,##0.00',
                    bg=data_bg, fg=num_fg, align="right"
                )
            elif ci == 1:
                _style_data(cell, bold=is_best, bg=data_bg, fg=num_fg, align="center")
            else:
                _style_data(cell, bold=is_best, bg=data_bg, fg=num_fg, align="left")
        ws.row_dimensions[row_num].height = 16

    note_row = 4 + len(quotes) + 2
    ws.merge_cells(f"A{note_row}:I{note_row}")
    note = ws[f"A{note_row}"]
    note.value = (
        "(*) Verde = mejor precio total. Rojo = precio más alto. "
        "Generado automáticamente por el sistema de gestión de cotizaciones."
    )
    note.font = Font(name="Calibri", italic=True, size=9, color="595959")
    note.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[note_row].height = 14

    ws.freeze_panes = ws["A4"]
    ws.sheet_view.showGridLines = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True


def export_to_excel(quotes: List[Quote]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _write_comparison_sheet(wb, quotes)
    _write_summary_sheet(wb, quotes)

    for q in quotes:
        _write_quote_sheet(wb, q)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
